# -*- coding: utf-8 -*-
"""
Сборка шаблона экспорта КП из финальной формы ВОР сметного отдела.

Что делает:
  1. Берёт финальную форму `temp/Пример выгрузки ВОР.xlsx` — листы КП / МАТЕРИАЛЫ / РАБОТЫ
     (колонка «Тип» в ней УЖЕ есть — вставлять её больше не нужно).
  2. Очищает ЗНАЧЕНИЯ динамической зоны листа «КП» (строки 18–45: таблица + ИТОГО + НДС) —
     стили сохраняет. Эти строки на экспорте перегенерирует серверный writer
     (server/src/lib/estimate-export).
  3. Очищает ЗНАЧЕНИЯ строк данных и строки-итога листов «МАТЕРИАЛЫ»/«РАБОТЫ» (с строки 4
     включительно до строки SUBTOTAL) — заголовок (строки 1–3) и стили строк-образцов
     сохраняются. Эти строки writer тоже перегенерирует.
  4. Статичный «хвост» листа «КП» (условия расценок, квалиф. блок) сохраняется.

Почему отдельным скриптом, а не в рантайме:
  Форма пересобирается РЕДКО — только когда сметный отдел присылает новую версию. Поэтому это
  офлайн-билдер: прогнать вручную, результат (бинарный .xlsx) закоммитить как ассет. openpyxl
  сохраняет стили/заливки/условное форматирование, значения динамической зоны — обнуляет.

Запуск (из корня репозитория):
  python scripts/estimate-export/build-kp-template.py
Результат: server/src/templates/kp-export-template.xlsx

Если отдел пришлёт новую форму — заменить SRC ниже на её путь и сверить раскладку с конфигом
writer'а (layout.ts в server/src/lib/estimate-export): имена листов, REF_COL, TABLE_START_ROW,
STYLE_ROW, TAIL_START_ROW, REF_SUBTOTAL_STYLE_ROW.
"""
import sys, io, os

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
def rel(*p): return os.path.join(REPO, *p)

# Источник задан ЯВНО (не glob): в temp/ лежит несколько *ВОР*.xlsx, и подбор по маске мог бы
# взять устаревший файл. Финальная форма отдела — «Пример выгрузки ВОР.xlsx».
SRC = rel('temp', 'Пример выгрузки ВОР.xlsx')
OUT = rel('server', 'src', 'templates', 'kp-export-template.xlsx')

if not os.path.exists(SRC):
    print('НЕ найден исходный файл формы:', SRC)
    sys.exit(1)

wb = openpyxl.load_workbook(SRC)
assert 'КП' in wb.sheetnames, f'в {os.path.basename(SRC)} нет листа «КП» (это не форма ВОР?)'
for sh in ('МАТЕРИАЛЫ', 'РАБОТЫ'):
    assert sh in wb.sheetnames, f'в {os.path.basename(SRC)} нет листа «{sh}»'
print('источник:', os.path.basename(SRC))

# 1) Лист «КП»: очистить ЗНАЧЕНИЯ динамической зоны, стили сохранить. Зона — от TABLE_START до
#    строки перед «хвостом»; хвост определяем по первой строке-условию (столбец A содержит «!!!»).
KP_TABLE_START = 18
ws = wb['КП']
tail_row = None
for r in range(KP_TABLE_START, ws.max_row + 1):
    a = ws.cell(row=r, column=1).value
    if isinstance(a, str) and a.lstrip().startswith('!!!'):
        tail_row = r
        break
assert tail_row is not None, 'не найдено начало «хвоста» (строка с «!!!») на листе КП'
# Между последней строкой ИТОГО/НДС и «хвостом» есть пустой спейсер — очищаем зону до него.
for r in range(KP_TABLE_START, tail_row):
    for c in range(1, ws.max_column + 1):
        ws.cell(row=r, column=c).value = None
print(f'КП: очищена зона строк {KP_TABLE_START}–{tail_row - 1} (хвост с {tail_row})')

# 2) Листы «МАТЕРИАЛЫ»/«РАБОТЫ»: очистить ЗНАЧЕНИЯ строк данных и строки-итога (с 4 до SUBTOTAL
#    включительно), заголовок (1–3) и стили сохранить.
REF_DATA_START = 4
for sh in ('МАТЕРИАЛЫ', 'РАБОТЫ'):
    w = wb[sh]
    sub_row = None
    for r in range(REF_DATA_START, w.max_row + 1):
        f = w.cell(row=r, column=6).value  # F — строка-итог содержит SUBTOTAL
        if isinstance(f, str) and 'SUBTOTAL' in f.upper():
            sub_row = r
            break
    assert sub_row is not None, f'на листе «{sh}» не найдена строка SUBTOTAL'
    for r in range(REF_DATA_START, sub_row + 1):
        for c in range(1, w.max_column + 1):
            w.cell(row=r, column=c).value = None
    print(f'{sh}: очищены строки данных {REF_DATA_START}–{sub_row} (итог был на {sub_row})')

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print('сохранено:', os.path.relpath(OUT, REPO))

# краткая проверка результата
v = openpyxl.load_workbook(OUT)
print('листы:', v.sheetnames)
kp = v['КП']
hdr = [kp.cell(row=15, column=c).value for c in range(1, 16)]
print('шапка(15):', ' | '.join(str(x) for x in hdr if x))
print('КП!C5:', kp['C5'].value, '| КП!C6:', kp['C6'].value, '(должны быть пусты — заполнит writer)')
