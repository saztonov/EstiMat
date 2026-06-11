"""Генерирует temp/справочник/Справочник_ЭОМ_v2.xlsx.
Листы: Работы | Материалы к работам | Материалы | Источники | Легенда
Запуск: python scripts/vor/gen_excel_v2.py  (из корня проекта)
Зависимость: pip install openpyxl
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT    = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SDIR    = os.path.join(ROOT, 'temp', 'справочник')
CAT_F   = os.path.join(SDIR, 'catalog_v2.json')
MAP_F   = os.path.join(SDIR, 'mapping_v2.json')
OUT_F   = os.path.join(SDIR, 'Справочник_ЭОМ_v2.xlsx')

with open(CAT_F, encoding='utf-8') as f: cat = json.load(f)
with open(MAP_F, encoding='utf-8') as f: mapping = json.load(f)
by_name = mapping['byName']
works   = cat['works']

BLUE  = PatternFill('solid', fgColor='1F497D')
LGRAY = PatternFill('solid', fgColor='D9D9D9')
YLLW  = PatternFill('solid', fgColor='FFEB9C')
GRN   = PatternFill('solid', fgColor='C6EFCE')
ORG   = PatternFill('solid', fgColor='FFCC99')
thin  = Border(left=Side(style='thin'), right=Side(style='thin'),
               top=Side(style='thin'),  bottom=Side(style='thin'))

def head(ws, cols):
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = BLUE
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin
    ws.row_dimensions[1].height = 30

def style_row(ws, ri, fill=None):
    for cell in ws[ri]:
        if fill: cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin

def widths(ws, ww):
    for i, w in enumerate(ww, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ── Лист 1: Работы ───────────────────────────────────────────────────────
ws1 = wb.active; ws1.title = 'Работы'
head(ws1, ['Категория','Вид (costType)','Типовая работа','Ед.',
           'Проектов','ВОР файлов','Алиасы (через ; )',
           'Расценка БД (имя)','Расценка БД (id)','Вид связи'])
widths(ws1, [28,38,50,8,9,9,60,60,38,10])

for w in works:
    m = by_name.get(w['name'], {})
    kind = m.get('kind','')
    rname, rid = m.get('rateName',''), m.get('rateId','')
    if kind == 'probable' and m.get('candidates'):
        rname = ' / '.join(c['rateName'] for c in m['candidates'])
        rid   = ' / '.join(c['rateId']   for c in m['candidates'])
    aliases = '; '.join(a['name'] for a in w.get('aliases',[]))
    ws1.append([w.get('category',''), w.get('costType',''), w['name'], w.get('unit',''),
                w.get('projectsCount',''), w.get('filesCount',''), aliases, rname, rid, kind])
    fill = GRN if kind=='matched' else (ORG if kind=='probable' else (LGRAY if kind=='toCreate' else None))
    style_row(ws1, ws1.max_row, fill)
ws1.freeze_panes = 'A2'

# ── Лист 2: Материалы к работам ──────────────────────────────────────────
ws2 = wb.create_sheet('Материалы к работам')
head(ws2, ['Работа','Вид (costType)','Материал','Ед.','Коэфф. расхода',
           'Проектов','ВОР файлов','Типовой'])
widths(ws2, [50,38,50,8,14,9,9,9])

for w in works:
    for mat in w.get('materials',[]):
        r = mat.get('ratioMedian')
        ws2.append([w['name'], w.get('costType',''), mat['name'], mat.get('unit',''),
                    f'{r:.4f}' if r is not None else 'нет данных',
                    mat.get('projectsCount',''), mat.get('filesCount',''),
                    'да' if mat.get('isTypical') else 'нет'])
        style_row(ws2, ws2.max_row, GRN if mat.get('isTypical') else None)
ws2.freeze_panes = 'A2'

# ── Лист 3: Материалы (уникальные) ───────────────────────────────────────
ws3 = wb.create_sheet('Материалы')
head(ws3, ['Материал','Ед.','Встречается в N работах','Типовой в N работах',
           'Макс. ВОР файлов','Макс. проектов'])
widths(ws3, [55,8,22,22,16,14])

stats = {}
for w in works:
    for mat in w.get('materials',[]):
        n = mat['name']
        if n not in stats:
            stats[n] = {'unit':mat.get('unit',''),'total':0,'typ':0,'files':0,'proj':0}
        s = stats[n]
        s['total'] += 1
        s['typ']   += 1 if mat.get('isTypical') else 0
        s['files']  = max(s['files'], mat.get('filesCount',0) or 0)
        s['proj']   = max(s['proj'],  mat.get('projectsCount',0) or 0)

for mn, s in sorted(stats.items(), key=lambda x: -x[1]['typ']):
    ws3.append([mn, s['unit'], s['total'], s['typ'], s['files'], s['proj']])
    style_row(ws3, ws3.max_row, GRN if s['typ'] > 0 else None)
ws3.freeze_panes = 'A2'

# ── Лист 4: Источники ─────────────────────────────────────────────────────
ws4 = wb.create_sheet('Источники')
head(ws4, ['Проект','Файл (имя)','Исключён как дубль'])
widths(ws4, [14,75,20])

DEDUP_KEY = 'ВОР_К6_ЭМ1-1.6_от_20.08.25_ЭО1-1.6_от_20.08.25'
file2proj = {}
for w in works:
    for a in w.get('aliases',[]):
        if a.get('file') and a.get('project'):
            file2proj[a['file']] = a['project']

for src in cat['meta'].get('sources',[]):
    exc = 'да' if src == DEDUP_KEY else ''
    ws4.append([file2proj.get(src,'?'), src, exc])
    style_row(ws4, ws4.max_row, LGRAY if exc else None)

note = cat['meta'].get('dedupNote','')
if note:
    ws4.append([])
    ws4.append(['Примечание о дублях:', note, ''])
    ri = ws4.max_row
    for cell in ws4[ri]:
        cell.fill = YLLW
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws4.row_dimensions[ri].height = 60
    ws4.merge_cells(f'B{ri}:C{ri}')
ws4.freeze_panes = 'A2'

# ── Лист 5: Легенда ───────────────────────────────────────────────────────
ws5 = wb.create_sheet('Легенда')
ws5.column_dimensions['A'].width = 14
ws5.column_dimensions['B'].width = 65
for label, desc, fill in [
    ('Зелёный',  'Точное совпадение с расценкой БД (matched) / Типовой материал', GRN),
    ('Оранжевый','Вероятное совпадение, требует подтверждения (probable)',         ORG),
    ('Серый',    'Нет аналога в БД, нужно создать (toCreate) / дубль-исходник',   LGRAY),
]:
    ws5.append([label, desc])
    for cell in ws5[ws5.max_row]:
        cell.fill = fill; cell.border = thin
        cell.alignment = Alignment(wrap_text=True, vertical='top')

wb.save(OUT_F)

n_mat  = sum(len(w.get('materials',[])) for w in works)
n_typ  = sum(1 for w in works for m in w.get('materials',[]) if m.get('isTypical'))
print(f'Excel: {OUT_F}')
print(f'  Работы: {len(works)}   Материалов: {n_mat} (типовых: {n_typ})   Уникальных мат.: {len(stats)}')
print(f'  Маппинг: {mapping["meta"]["matched"]} matched / {mapping["meta"]["probable"]} probable / {mapping["meta"]["toCreate"]} toCreate')
